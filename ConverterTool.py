# importing the module
import json
import openpyxl
import string
import os
import re
import sys, getopt
from pathlib import Path
import time

# Setting the path to the xlsx file:
try:
    opt, args = getopt.gnu_getopt(sys.argv[1:], "ot:r:", longopts=['old', 'template=', 'resultPath='])

    xlsx_file = os.path.join(args[0]) #os.path.join('C:\\x\\tools', 'Elienai.xlsm') 
    name = args[1]
    file_name = name.replace(" ", "_")+".json"
    targetFile = file_name
    templateFile = 'actorSmart.json'
    actor = {
        "name": name,
        "type": "character"
    }
    for o, a in opt:
        if o in ["-t", "--template"]:
            templateFile = a
        elif o in ["-r", "--resultPath"]:
            if(not os.path.exists(a)):
                os.makedirs(a)
            targetFile = os.path.join(a, file_name)
       

    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
except Exception as e:
    print(e)
    time.sleep(15)

def getCell(cell):
    sheet = wb.active
    return sheet[cell].value

def getValue(name):
    if ('$' in name):
        cell = name.split('$')
        if not cell[0]: 
            cell[0] = 'Principal'
        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == cell[0]:
                break
        wb.active = s
        return getCell(cell[1])
    # get DefinedNameList instance
    defined_name = wb.defined_names[name]

    # get destinations which is a generator
    destinations = defined_name.destinations  
    values = {}
    for sheet_title, coordinate in destinations:
        values.update({sheet_title: wb[sheet_title][coordinate].value})
    my_value = values.get(sheet_title)
    return my_value

def fillDict(data):
    for key in data:
        if type(data[key]) is dict:
            #If field is a dictionary, fill said dictionary
            data[key] = fillDict(data[key])
        elif type(data[key]) is str and data[key]:
            #If element is a non empty String, swap it by a value from th Excel
            data[key] = getValue(data[key])
    return data
    
try:
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Principal':
            break
    wb.active = s
    principal = wb.active
    with open(templateFile) as json_file:
        data = json.load(json_file)

        secundariasEspeciales = {}
        for i in range(73, 77):
            valor = principal.cell(row=i, column=17).value
            if valor != "-":
                secondary = {"base":{"value":valor}, "final":{"value": 0}}
                secundariasEspeciales .update({principal.cell(row=i, column=13).value: secondary})
        data["secondaries"]["secondarySpecialSkills"] = secundariasEspeciales

        for s in range(len(wb.sheetnames)):
            if wb.sheetnames[s] == 'Místicos':
                break
        wb.active = s
        magic = wb.active

        #TODO Imbalance

        for path in data["mystic"]["magicLevel"]["spheres"]:
            via = data["mystic"]["magicLevel"]["spheres"][path]["value"]
            for i in range(15, 25):
                if magic.cell(row=i, column=3).value == via:
                    data["mystic"]["magicLevel"]["spheres"][path]["value"] = magic.cell(row=i, column=8).value
            if data["mystic"]["magicLevel"]["spheres"][path]["value"] == via:
                data["mystic"]["magicLevel"]["spheres"][path]["value"] = 0

        #data["mystic"]["magicLevel"]["used"]["value"] = 

        data = fillDict(data)


        actor.update({"data": data})


    with open(targetFile, 'w') as outfile:
        json.dump(actor, outfile)
except Exception as e:
    print(e)
    time.sleep(15)



    

